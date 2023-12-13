import pymysql
import openpyxl
import tkinter as tk
from tkinter import scrolledtext


def execute_queries_and_export_results(salon_id):
    # Устанавливаем подключение к базе данных (Добавь доступы)
    connection = pymysql.connect(
        host='',
        user='',
        password='',
        database=''
    )

    try:
        # Создаем новый Excel-файл
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Устанавливаем заголовки столбцов
        sheet.cell(row=1, column=1).value = 'Date' #0
        sheet.cell(row=2, column=1).value = 'Онлайн-Запись' #1
        sheet.cell(row=3, column=1).value = 'Использование лояльности' #2
        sheet.cell(row=4, column=1).value = 'Расчет зарплат' #3
        sheet.cell(row=5, column=1).value = 'Печать чеков' #4
        sheet.cell(row=6, column=1).value = 'Складские операции' #5
        sheet.cell(row=7, column=1).value = 'Финансовые операции' #6
        sheet.cell(row=8, column=1).value = 'офлайн-записи'  # 6
        sheet.cell(row=9, column=1).value = 'События' #9
        #Офлайн записи 7
        #Онлайн оплата записей 8

        # Устанавливаем значения для столбца Date
        dates = [0, 1, 2, 3, 4, 5]  # Список значений для столбца Date
        with connection.cursor() as cursor:
            for i, date in enumerate(dates):
                query0 = f"select CONCAT(month(DATE_ADD(CURRENT_DATE,interval -{date} month)), '-', year(CURRENT_DATE)) "
                cursor.execute(query0)
                result = cursor.fetchone()
                sheet.cell(row=1, column=i+2).value = result[0]

        # Выполняем запросы и записываем результаты в ячейки
        row = 2  # Начальная строка для результатов
        with connection.cursor() as cursor:
            for date in dates:
                # Выполняем запрос для loyalty_transactions
                query1 = f"SELECT COUNT(id) FROM loyalty_transactions " \
                         f"WHERE month(created) = month(DATE_ADD(CURRENT_DATE,interval -{date} month)) " \
                         f"AND year(created) = year(CURRENT_DATE) " \
                         f"and salon_id = {salon_id}"
                cursor.execute(query1)
                result1 = cursor.fetchone()
                sheet.cell(row=3, column=row).value = result1[0]

                # Выполняем запрос для transactions
                query2 = f"SELECT COUNT(id) FROM transactions " \
                         f"WHERE month(date_create) = month(DATE_ADD(CURRENT_DATE,interval -{date} month)) " \
                         f"AND year(date_create) = year(CURRENT_DATE) " \
                         f"and account_id in (select id from accounts where salon_id = {salon_id})"
                cursor.execute(query2)
                result2 = cursor.fetchone()
                sheet.cell(row=7, column=row).value = result2[0]

                query3 = f"select count(id) from tt_records where type in (11,2,10,3,1,6,8,9) AND (bookform_id != 0 or partner_id != 0) " \
                         f"AND (month(cdate) = month(DATE_ADD(CURRENT_DATE,interval -{date} month)) AND year(cdate) = year(CURRENT_DATE)) " \
                         f"AND salon_id = {salon_id}"
                cursor.execute(query3)
                result3 = cursor.fetchone()
                sheet.cell(row=2, column=row).value = result3[0]

                query4 = f"select count(id) from activities " \
                         f"where salon_id = {salon_id} " \
                         f"AND month(created) = month(DATE_ADD(CURRENT_DATE,interval -{date} month)) " \
                         f"AND year(created) = year(CURRENT_DATE)"
                cursor.execute(query4)
                result4 = cursor.fetchone()
                sheet.cell(row=9, column=row).value = result4[0]

                query5 = f"select count(id) from goods_transactions where salon_id = {salon_id} " \
                         f"AND month(create_date_original) = month(DATE_ADD(CURRENT_DATE,interval -{date} month)) " \
                         f"AND year(create_date_original) = year(CURRENT_DATE)"
                cursor.execute(query5)
                result5 = cursor.fetchone()
                sheet.cell(row=6, column=row).value = result5[0]

                query6 = f"select count(id) from kkm_transactions " \
                         f"where salon_id = {salon_id} " \
                         f"AND month(print_date) = month(DATE_ADD(CURRENT_DATE,interval -{date} month)) " \
                         f"AND year(print_date) = year(CURRENT_DATE)"
                cursor.execute(query6)
                result6 = cursor.fetchone()
                sheet.cell(row=5, column=row).value = result6[0]

                query7 = f"select count(id) from master_salary_calculation " \
                         f"where salon_id = {salon_id} " \
                         f"AND month(date) = month(DATE_ADD(CURRENT_DATE,interval -{date} month)) " \
                         f"AND year(date) = year(CURRENT_DATE)"
                cursor.execute(query7)
                result7 = cursor.fetchone()
                sheet.cell(row=4, column=row).value = result7[0]

                query8 = f"select count(id) from tt_records where type in (0, 12) " \
                         f"AND (month(cdate) = month(DATE_ADD(CURRENT_DATE,interval -{date} month)) AND year(cdate) = year(CURRENT_DATE)) " \
                         f"AND salon_id = {salon_id}"
                cursor.execute(query8)
                result8 = cursor.fetchone()
                sheet.cell(row=8, column=row).value = result8[0]

                row += 1  # Переходим к следующей строке для результатов
                set_column_width(sheet)
        # Сохраняем файл Excel
        workbook.save(f'филиал_{salon_id}.xlsx')
        console_text.insert(tk.END, 'Результаты успешно записаны в файл results.xlsx\n')

    except Exception as err:
        console_text.insert(tk.END, f"Ошибка: {err}")

    finally:
        # Закрываем соединение с базой данных
        connection.close()

def set_column_width(sheet):
    # Получаем максимальную длину текста в каждом столбце
    max_length = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell.column_letter not in max_length:
                    max_length[cell.column_letter] = cell_length
                else:
                    if cell_length > max_length[cell.column_letter]:
                        max_length[cell.column_letter] = cell_length
    # Устанавливаем ширину столбца в соответствии с максимальной длиной текста
    for column, width in max_length.items():
        sheet.column_dimensions[column].width = width


# Вызываем функцию для выполнения запросов и экспорта результатов
def process_input():
    salon_id = input_entry.get()
    execute_queries_and_export_results(salon_id)
    input_entry.delete(0, tk.END)


window = tk.Tk()
window.title("Аналитика использования модулей")

input_entry = tk.Label(window, text="Введи ID филиала")
input_entry.pack()
input_entry = tk.Entry(window)
input_entry.pack()

process_button = tk.Button(window, text="Собрать отчет", command=process_input)
process_button.pack()

console_text = scrolledtext.ScrolledText(window)
console_text.configure(bg="#1a4780", fg="yellow")
console_text.pack()
window_width = 320
window_height = 240
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
x_coordinate = (screen_width - window_width) // 2
y_coordinate = (screen_height - window_height) // 2
window.geometry(f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")



# Запускаем главный цикл обработки событий Tkinter
window.mainloop()
